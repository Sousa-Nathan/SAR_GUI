""" Compares changes between data sheets of identical size """

import pandas as pd
import numpy as np
import openpyxl
import logging
import os
from openpyxl.styles import (Font, PatternFill)
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle

class Compare:

    def __init__(self, sum_workbook_1: str, sum_workbook_2: str, comparison_filepath: str, log_dir: str):
        self.sw1 = sum_workbook_1
        self.sw2 = sum_workbook_2
        self.cf = comparison_filepath
        self.log = log_dir

    def compare_workbooks(self):
        try:
            df_wb1 = pd.read_excel(self.sw1, sheet_name = "Worst Case SAR")
            df_wb2 = pd.read_excel(self.sw2, sheet_name = "Worst Case SAR")
            
            df_wb1.equals(df_wb2)
            
            comparison_values = df_wb1.values == df_wb2.values
            
            rows, cols = np.where(comparison_values == False)
            
            # with open(f"{self.log}/rows_cols.txt", "w") as f:
            #     for item in zip(rows, cols):
            #         f.write(f"WB1 R{item[0]}:C{item[1]} - Type: {type(df_wb1.iloc[item[0], item[1]])} - Value: {df_wb1.iloc[item[0], item[1]]} | WB2 R{item[0]}:C{item[1]} - Type: {type(df_wb2.iloc[item[0], item[1]])} - Value: {df_wb2.iloc[item[0], item[1]]}\n")
            
            for item in zip(rows, cols):
                if df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 0]:
                    df_wb1.iloc[item[0], item[1]] = f"New: {df_wb2.iloc[item[0], 0]} | Delta: {df_wb1.iloc[item[0], 0]} -> {df_wb2.iloc[item[0], 0]}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 1]:
                    df_wb1.iloc[item[0], item[1]] = f"New: {df_wb2.iloc[item[0], 1]} | Delta: {df_wb1.iloc[item[0], 1]} -> {df_wb2.iloc[item[0], 1]}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 2]:
                    df_wb1.iloc[item[0], item[1]] = f"New: {df_wb2.iloc[item[0], 2]} | Delta: {df_wb1.iloc[item[0], 2]} -> {df_wb2.iloc[item[0], 2]}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 3]:
                    df_wb1.iloc[item[0], item[1]] = f"New: {df_wb2.iloc[item[0], 3]} | Delta: {df_wb1.iloc[item[0], 3]} -> {df_wb2.iloc[item[0], 3]}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 4]:
                    df_wb1.iloc[item[0], item[1]] = f"New: {df_wb2.iloc[item[0], 4]} | Delta: {df_wb1.iloc[item[0], 4]} -> {df_wb2.iloc[item[0], 4]}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 12]:
                    df_wb1.iloc[item[0], item[1]] = f"New: {df_wb2.iloc[item[0], 12]} | Delta: {df_wb1.iloc[item[0], 12]} -> {df_wb2.iloc[item[0], 12]}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 13]:
                    df_wb1.iloc[item[0], item[1]] = f"New: {df_wb2.iloc[item[0], 13]} | Delta: {df_wb1.iloc[item[0], 13]} -> {df_wb2.iloc[item[0], 13]}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 14]:
                    df_wb1.iloc[item[0], item[1]] = f"New: {df_wb2.iloc[item[0], 14]} | Delta: {df_wb1.iloc[item[0], 14]} -> {df_wb2.iloc[item[0], 14]}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 15]:
                    wb1 = f"{df_wb1.iloc[item[0], 15]:.0f}"
                    wb2 = f"{df_wb2.iloc[item[0], 15]:.0f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 16]:
                    wb1 = f"{df_wb1.iloc[item[0], 16]:.0f}"
                    wb2 = f"{df_wb2.iloc[item[0], 16]:.0f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 17]:
                    wb1 = f"{df_wb1.iloc[item[0], 17]:.2%}"
                    wb2 = f"{df_wb2.iloc[item[0], 17]:.2%}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 18]:
                    wb1 = f"{df_wb1.iloc[item[0], 18]:.3f}"
                    wb2 = f"{df_wb2.iloc[item[0], 18]:.3f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 19]:
                    wb1 = f"{df_wb1.iloc[item[0], 19]:.2f}"
                    wb2 = f"{df_wb2.iloc[item[0], 19]:.2f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 20]:
                    wb1 = f"{df_wb1.iloc[item[0], 20]:.2f}"
                    wb2 = f"{df_wb2.iloc[item[0], 20]:.2f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 21]:
                    wb1 = f"{df_wb1.iloc[item[0], 21]:.3f}"
                    wb2 = f"{df_wb2.iloc[item[0], 21]:.3f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 22]:
                    wb1 = f"{df_wb1.iloc[item[0], 22]:.3f}"
                    wb2 = f"{df_wb2.iloc[item[0], 22]:.3f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 23]:
                    wb1 = f"{df_wb1.iloc[item[0], 23]:.3f}"
                    wb2 = f"{df_wb2.iloc[item[0], 23]:.3f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 24]:
                    wb1 = f"{df_wb1.iloc[item[0], 24]:.3f}"
                    wb2 = f"{df_wb2.iloc[item[0], 24]:.3f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 25]:
                    wb1 = f"{df_wb1.iloc[item[0], 25]:.3f}"
                    wb2 = f"{df_wb2.iloc[item[0], 25]:.3f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 26]:
                    wb1 = f"{df_wb1.iloc[item[0], 26]:.3f}"
                    wb2 = f"{df_wb2.iloc[item[0], 26]:.3f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 27]:
                    wb1 = f"{df_wb1.iloc[item[0], 27]:.3f}"
                    wb2 = f"{df_wb2.iloc[item[0], 27]:.3f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 28]:
                    wb1 = f"{df_wb1.iloc[item[0], 28]:.3f}"
                    wb2 = f"{df_wb2.iloc[item[0], 28]:.3f}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
                elif df_wb1.iloc[item[0], item[1]] == df_wb1.iloc[item[0], 29]:
                    wb1 = f"{df_wb1.iloc[item[0], 29]:.2%}"
                    wb2 = f"{df_wb2.iloc[item[0], 29]:.2%}"
                    
                    df_wb1.iloc[item[0], item[1]] = f"New: {wb2} | Delta: {wb1} -> {wb2}"
            
            for item in zip(rows, cols):
                if df_wb1.iloc[item[0], item[1]] == "New: nan | Delta: nan -> nan" or df_wb1.iloc[item[0], item[1]] == "New: .3f | Delta: .3f -> .3f":
                    df_wb1.iloc[item[0], item[1]] = ""
            
            df_wb1.to_excel(f"{self.cf}", sheet_name = "Summary Comparison", index = False, header = True)
        
        except Exception as e:
            logger = logging.getLogger("SAR GUI: COMPARE Module")
            logger.setLevel(logging.ERROR)
            
            lfh = logging.FileHandler(os.path.join(self.log, "error.log"))
            lfh.setLevel(logging.ERROR)
            
            formatter = logging.Formatter("\n%(asctime)s - %(name)s - %(levelname)s - %(message)s")
            lfh.setFormatter(formatter)
            
            logger.addHandler(lfh)
            
            logger.exception(e)

    def compare_format_results(self):
        wb = openpyxl.load_workbook(f"{self.cf}")
        
        ws = wb["Summary Comparison"]
        
        red_font = Font(color="FF0000")
        yellow_fill = PatternFill(fill_type = "solid",
                                  bgColor = "00FFFF00")
        dxf = DifferentialStyle(font = red_font,
                                fill = yellow_fill)
        
        rule = Rule(type = "containsText", operator = "containsText", text = "Delta", dxf=dxf)
        rule.formula = ['NOT(ISERROR(SEARCH("Delta",A2)))']
        
        ws.conditional_formatting.add("A2:AD15000", rule)
        
        wb.save(f"{self.cf}")
