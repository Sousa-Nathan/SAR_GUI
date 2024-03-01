"""
GUI to compile raw data ready for reports, generate
a summary workbook, and produce variant test plans.
"""

import logging
import os
import random
from datetime import datetime

import openpyxl as op
import pandas as pd
import PySimpleGUI as sg
from Modules.COMPARE import Compare
from Modules.Not_WiFi import Not_WiFi
from Modules.SEC1 import Sec_1
from Modules.WLAN import Wlan
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Side)

def parse_reported_workbook_name(sprsht_fp):
    """
    Parse the name of the workbook to be generated from the workbook file path.
    
    :param sprsht_fp: excel workbook file path
    :return: workbook name, 'workbook_name.xlsx'
    """
    
    # directory_names = sprsht_fp.split("/")
    now = datetime.now()
    date = now.strftime("%Y_%m_%d")
    time = now.strftime("%H_%M_%S")
    reported_workbook_name = f"SAR_Date-{date}_Time-{time}.xlsx"
    
    return reported_workbook_name

def parse_summary_workbook_name(sprsht_fp):
    """
    Parse the name of the workbook to be generated from the workbook file path.
    
    :param sprsht_fp: excel workbook file path
    :return: workbook name, 'workbook_name.xlsx'
    """
    
    # directory_names = sprsht_fp.split("/")
    now = datetime.now()
    date = now.strftime("%Y_%m_%d")
    time = now.strftime("%H_%M_%S")
    summary_workbook_name = f"Worst_Case_Summary_SAR_Date-{date}_Time-{time}.xlsx"
    
    return summary_workbook_name

def parse_smtx_workbook_name(sprsht_fp):
    """
    Parse the name of the workbook to be generated from the workbook file path.
    
    :param sprsht_fp: excel workbook file path
    :return: workbook name, 'workbook_name.xlsx'
    """
    
    # directory_names = sprsht_fp.split("/")
    now = datetime.now()
    date = now.strftime("%Y_%m_%d")
    time = now.strftime("%H_%M_%S")
    smtx_workbook_name = f"SmartTx_SAR_Data_Date-{date}_Time-{time}.xlsx"
    
    return smtx_workbook_name

def parse_sec1_fcc_workbook_name(sprsht_fp):
    """
    Parse the name of the workbook to be generated from the workbook file path.
    
    :param sprsht_fp: excel workbook file path
    :return: workbook name, 'workbook_name.xlsx'
    """
    
    # directory_names = sprsht_fp.split("/")
    now = datetime.now()
    date = now.strftime("%Y_%m_%d")
    time = now.strftime("%H_%M_%S")
    sec1_fcc_workbook_name = f"Sec_1_FCC_Date-{date}_Time-{time}.xlsx"
    
    return sec1_fcc_workbook_name

def parse_sec1_ised_workbook_name(sprsht_fp):
    """
    Parse the name of the workbook to be generated from the workbook file path.
    
    :param sprsht_fp: excel workbook file path
    :return: workbook name, 'workbook_name.xlsx'
    """
    
    # directory_names = sprsht_fp.split("/")
    now = datetime.now()
    date = now.strftime("%Y_%m_%d")
    time = now.strftime("%H_%M_%S")
    sec1_ised_workbook_name = f"Sec_1_ISED_Date-{date}_Time-{time}.xlsx"
    
    return sec1_ised_workbook_name

def parse_summary_comparison_workbook_name(sprsht_fp):
    """
    Parse the name of the workbook to be generated from the workbook file path.
    
    :param sprsht_fp: excel workbook file path
    :return: workbook name, 'workbook_name.xlsx'
    """
    
    # directory_names = sprsht_fp.split("/")
    now = datetime.now()
    date = now.strftime("%Y_%m_%d")
    time = now.strftime("%H_%M_%S")
    sum_comparison_workbook_name = f"Summary_Comparison_Date-{date}_Time-{time}.xlsx"
    
    return sum_comparison_workbook_name

def reported_excel(filepath):
    """
    Formats the workbook for readability.

    Args:
        filepath (str): String filepath to the raw workbook
    """
    
    workbook = op.load_workbook(filename = filepath)
    
    get_sheet = workbook.sheetnames
    
    header = NamedStyle(name = "header")
    header.font = Font(name = "Arial", sz = 8, bold = True)
    header.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    header.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    header.fill = PatternFill(fill_type = "solid", start_color = "00538DD5")
    
    formatted_cells = NamedStyle(name = "formatted_cells")
    formatted_cells.font = Font(name = "Arial", sz = 8)
    formatted_cells.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    formatted_cells.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    
    for sheet in get_sheet:
        active_sheet = workbook[sheet]
        dimensions = active_sheet.dimensions
        header_row = active_sheet[1]
        
        active_sheet.column_dimensions["A"].width = 13.15
        active_sheet.column_dimensions["B"].width = 12.72
        active_sheet.column_dimensions["C"].width = 17.47
        active_sheet.column_dimensions["D"].width = 12.86
        active_sheet.column_dimensions["E"].width = 5.72
        active_sheet.column_dimensions["F"].width = 9.58
        active_sheet.column_dimensions["G"].width = 8.72
        active_sheet.column_dimensions["H"].width = 8.72
        active_sheet.column_dimensions["I"].width = 8.72
        active_sheet.column_dimensions["J"].width = 8.72
        active_sheet.column_dimensions["K"].width = 9.15
        active_sheet.column_dimensions["L"].width = 13.22
        active_sheet.column_dimensions["M"].width = 10.29
        active_sheet.column_dimensions["N"].width = 10.29
        active_sheet.column_dimensions["O"].width = 10.72
        active_sheet.column_dimensions["P"].width = 10.72
        active_sheet.column_dimensions["Q"].width = 10.72
        active_sheet.column_dimensions["R"].width = 10.72
        active_sheet.column_dimensions["S"].width = 10.72
        active_sheet.column_dimensions["T"].width = 10.72
        active_sheet.column_dimensions["U"].width = 10.72
        active_sheet.column_dimensions["V"].width = 10.72
        active_sheet.column_dimensions["W"].width = 15.58
        active_sheet.column_dimensions["X"].width = 10.72
        active_sheet.column_dimensions["Y"].width = 4.72
        
        for row in active_sheet[f"{dimensions}"]:
            for cell in row:
                cell.style = formatted_cells
                cell.style = formatted_cells
        
        for row in active_sheet["K2:K1000"]:
            for cell in row:
                cell.number_format = "0.00%"
        
        for row in active_sheet["L2:L1000"]:
            for cell in row:
                cell.number_format = "0.000"
        
        wifi_list = ['Wi-Fi 2.4 GHz', 'Wi-Fi 5.2 GHz', 'Wi-Fi 5.3 GHz', 'Wi-Fi 5.5 GHz', 'Wi-Fi 5.8 GHz', 'U-NII 5', 'U-NII 6', 'U-NII 7', 'U-NII 8']
        if any(sheet == active_sheet for sheet in wifi_list):
            for row in active_sheet["M2:N1000"]:
                for cell in row:
                    cell.number_format = "0.00"
        else:
            for row in active_sheet["M2:N1000"]:
                for cell in row:
                    cell.number_format = "0.0"
                    
        for row in active_sheet["O2:V1000"]:
            for cell in row:
                cell.number_format = "0.000"
                
        for row in active_sheet["X2:X1000"]:
            for cell in row:
                cell.number_format = "0.00"
                
        for cell in header_row:
            cell.style = header
    
    new_sheet = workbook.create_sheet("Author")
    
    author_list = ["Abhiyan Sapkota", "AJ Newcomer", "Alexa Sanchez Ochoa", "Anmol Ahmad", "Armando Gonzalez Hernandez", "Brandon Sousa", "Chris Kuwatani", "Coltyce Sanders", "David Cervantes", "David Weaver", "Devin Chang", "Jennifer Bunnell", "John Moestopo", "Kiara Davis", "Miguel Llamas", "Monika Sipa-Skorka", "Nina Villar", "Prem Dhar", "Remi Rodberg", "Richard Jankovics", "Ruben Lozano", "Samuel Yarman", "Tony Soares", "Truc Tran", "Ysidra van Kempen", "Zachary A. Coustier", "Zoe Moestopo"]
    
    new_sheet["A1"] = f"Brought to you by: {author_list[random.randrange(0, len(author_list))]}"
    new_sheet["A1"].font = Font(name = "Arial", sz = 72)
    
    workbook.save(filename = filepath)

def summary_excel(filepath):
    """
    Formats the workbook for readability.

    Args:
        filepath (str): String filepath to the raw workbook
    """
    
    workbook = op.load_workbook(filename = filepath)
    
    filter_page = workbook.active
    
    dimensions = filter_page.dimensions
    
    filter_page.freeze_panes = "A2"
    
    filter_page.auto_filter.ref = dimensions
    
    filter_page.column_dimensions["A"].width = 15.72
    filter_page.column_dimensions["B"].width = 13.01
    filter_page.column_dimensions["C"].width = 16.15
    filter_page.column_dimensions["D"].width = 15.58
    filter_page.column_dimensions["E"].width = 14.29
    filter_page.column_dimensions["F"].width = 12.15
    filter_page.column_dimensions["G"].width = 14.72
    filter_page.column_dimensions["H"].width = 13.72
    filter_page.column_dimensions["I"].width = 15.29
    filter_page.column_dimensions["J"].width = 17.43
    filter_page.column_dimensions["K"].width = 15.72
    filter_page.column_dimensions["L"].width = 15.72
    filter_page.column_dimensions["M"].width = 16.01
    filter_page.column_dimensions["N"].width = 12.01
    filter_page.column_dimensions["O"].width = 14.43
    filter_page.column_dimensions["P"].width = 15.72
    filter_page.column_dimensions["Q"].width = 16.58
    filter_page.column_dimensions["R"].width = 15.72
    filter_page.column_dimensions["S"].width = 10.29
    filter_page.column_dimensions["T"].width = 10.29
    filter_page.column_dimensions["U"].width = 10.29
    filter_page.column_dimensions["V"].width = 13.29
    filter_page.column_dimensions["W"].width = 13.29
    filter_page.column_dimensions["X"].width = 13.29
    filter_page.column_dimensions["Y"].width = 13.29
    filter_page.column_dimensions["Z"].width = 13.29
    filter_page.column_dimensions["AA"].width = 13.29
    filter_page.column_dimensions["AB"].width = 13.29
    filter_page.column_dimensions["AC"].width = 13.29
    filter_page.column_dimensions["AD"].width = 13.01
    
    header = NamedStyle(name = "header")
    header.font = Font(name = "Arial", sz = 8, bold = True)
    header.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    header.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    header.fill = PatternFill(fill_type = "solid", start_color = "00538DD5")
    
    header_row = filter_page[1]
    
    formatted_cells = NamedStyle(name = "formatted_cells")
    formatted_cells.font = Font(name = "Arial", sz = 8)
    formatted_cells.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    formatted_cells.alignment = Alignment(horizontal = "center", vertical = "center")
    
    for row in filter_page[f"{dimensions}"]:
        for cell in row:
            cell.style = formatted_cells
            cell.style = formatted_cells
    
    for row in filter_page["Q2:Q3000"]:
        for cell in row:
            cell.number_format = "0.00%"
    
    for row in filter_page["R2:R3000"]:
        for cell in row:
            cell.number_format = "0.000"
    
    for row in filter_page["S2:T3000"]:
        for cell in row:
            cell.number_format = "0.00"
    
    for row in filter_page["U2:AB3000"]:
        for cell in row:
            cell.number_format = "0.000"
    
    # for row in filter_page["AD2:AD3000"]:
    #     for cell in row:
    #         cell.number_format = "0.00%"
    
    for cell in header_row:
        cell.style = header
    
    new_sheet = workbook.create_sheet("Author")
    
    author_list = ["Abhiyan Sapkota", "AJ Newcomer", "Alexa Sanchez Ochoa", "Anmol Ahmad", "Armando Gonzalez Hernandez", "Brandon Sousa", "Chris Kuwatani", "Coltyce Sanders", "David Cervantes", "David Weaver", "Devin Chang", "Jennifer Bunnell", "John Moestopo", "Kiara Davis", "Miguel Llamas", "Monika Sipa-Skorka", "Nina Villar", "Prem Dhar", "Remi Rodberg", "Richard Jankovics", "Ruben Lozano", "Samuel Yarman", "Tony Soares", "Truc Tran", "Ysidra van Kempen", "Zachary A. Coustier", "Zoe Moestopo"]
    
    new_sheet["A1"] = f"Brought to you by: {author_list[random.randrange(0, len(author_list))]}"
    new_sheet["A1"].font = Font(name = "Arial", sz = 72)
    
    workbook.save(filename = filepath)

def summary_compare_excel(filepath, scwb1, scwb2):
    """
    Formats the workbook for readability and highlights cells that changed.
    
    Args:
        filepath (str): String filepath to the raw workbook
        scwb1 (dataframe): Data from one workbook
        scwb2 (dataframe): Data from one workbook
    """
    
    workbook = op.load_workbook(filename = filepath)
    
    filter_page = workbook.active
    
    dimensions = filter_page.dimensions
    
    filter_page.freeze_panes = "A2"
    
    filter_page.auto_filter.ref = dimensions
    
    filter_page.column_dimensions["A"].width = 37.72
    filter_page.column_dimensions["B"].width = 37.72
    filter_page.column_dimensions["C"].width = 70.72
    filter_page.column_dimensions["D"].width = 29.29
    filter_page.column_dimensions["E"].width = 40.29
    filter_page.column_dimensions["F"].width = 12.15
    filter_page.column_dimensions["G"].width = 14.72
    filter_page.column_dimensions["H"].width = 13.72
    filter_page.column_dimensions["I"].width = 15.29
    filter_page.column_dimensions["J"].width = 17.43
    filter_page.column_dimensions["K"].width = 15.72
    filter_page.column_dimensions["L"].width = 15.72
    filter_page.column_dimensions["M"].width = 36.86
    filter_page.column_dimensions["N"].width = 41.72
    filter_page.column_dimensions["O"].width = 40.01
    filter_page.column_dimensions["P"].width = 28.43
    filter_page.column_dimensions["Q"].width = 28.43
    filter_page.column_dimensions["R"].width = 29.29
    filter_page.column_dimensions["S"].width = 26.86
    filter_page.column_dimensions["T"].width = 26.72
    filter_page.column_dimensions["U"].width = 26.72
    filter_page.column_dimensions["V"].width = 24.86
    filter_page.column_dimensions["W"].width = 24.86
    filter_page.column_dimensions["X"].width = 24.86
    filter_page.column_dimensions["Y"].width = 24.86
    filter_page.column_dimensions["Z"].width = 24.86
    filter_page.column_dimensions["AA"].width = 24.86
    filter_page.column_dimensions["AB"].width = 24.86
    filter_page.column_dimensions["AC"].width = 24.86
    filter_page.column_dimensions["AD"].width = 29.29
    
    header = NamedStyle(name = "header")
    header.font = Font(name = "Arial", sz = 8, bold = True)
    header.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    header.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    header.fill = PatternFill(fill_type = "solid", start_color = "00538DD5")
    
    header_row = filter_page[1]
    
    formatted_cells = NamedStyle(name = "formatted_cells")
    formatted_cells.font = Font(name = "Arial", sz = 8)
    formatted_cells.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    formatted_cells.alignment = Alignment(horizontal = "center", vertical = "center")
    
    for row in filter_page[f"{dimensions}"]:
        for cell in row:
            cell.style = formatted_cells
            cell.style = formatted_cells
    
    for row in filter_page["R2:R3000"]:
        for cell in row:
            cell.number_format = "0.00%"
    
    for row in filter_page["S2:S3000"]:
        for cell in row:
            cell.number_format = "0.000"
    
    for row in filter_page["T2:U3000"]:
        for cell in row:
            cell.number_format = "0.00"
    
    for cell in header_row:
        cell.style = header
    
    source_workbooks = workbook.create_sheet("Source Workbooks")
    
    source_workbooks["A1"] = f"New: {scwb2} | Delta: {scwb1} -> {scwb2}"
    source_workbooks["A1"].font = Font(name = "Arial", bold = True, sz = 12)
    
    new_sheet = workbook.create_sheet("Author")
    
    author_list = ["Abhiyan Sapkota", "AJ Newcomer", "Alexa Sanchez Ochoa", "Anmol Ahmad", "Armando Gonzalez Hernandez", "Brandon Sousa", "Chris Kuwatani", "Coltyce Sanders", "David Cervantes", "David Weaver", "Devin Chang", "Jennifer Bunnell", "John Moestopo", "Kiara Davis", "Miguel Llamas", "Monika Sipa-Skorka", "Nina Villar", "Prem Dhar", "Remi Rodberg", "Richard Jankovics", "Ruben Lozano", "Samuel Yarman", "Tony Soares", "Truc Tran", "Ysidra van Kempen", "Zachary A. Coustier", "Zoe Moestopo"]
    
    new_sheet["A1"] = f"Brought to you by: {author_list[random.randrange(0, len(author_list))]}"
    new_sheet["A1"].font = Font(name = "Arial", sz = 72)
    
    workbook.save(filename = filepath)

def sec1_excel(filepath):
    """
    Formats the workbook for readability and generates variant test plans.
    
    Args:
        filepath (str): String filepath to the raw workbook
    """
    
    workbook = op.load_workbook(filename = filepath)
    
    get_sheet = workbook.sheetnames
    
    header = NamedStyle(name = "header")
    header.font = Font(name = "Arial", sz = 8, bold = True)
    header.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    header.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    header.fill = PatternFill(fill_type = "solid", start_color = "00538DD5")
    
    header_sec1 = NamedStyle(name = "header_sec1")
    header_sec1.font = Font(name = "Arial", sz = 8, bold = True)
    header_sec1.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    header_sec1.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    header_sec1.fill = PatternFill(fill_type = "solid", start_color = "00538DD5")
    header_sec1.fill = PatternFill(fill_type = "solid", start_color = "00C4BD97")
    
    formatted_cells = NamedStyle(name = "formatted_cells")
    formatted_cells.font = Font(name = "Arial", sz = 8)
    formatted_cells.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    formatted_cells.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    
    for sheet in get_sheet:
        active_sheet = workbook[sheet]
        dimensions = active_sheet.dimensions
        header_row = active_sheet[1]
        
        if active_sheet == workbook[get_sheet[0]]:
            active_sheet.column_dimensions["A"].width = 12.29
            active_sheet.column_dimensions["B"].width = 12.29
            active_sheet.column_dimensions["C"].width = 12.29
            active_sheet.column_dimensions["D"].width = 12.29
            active_sheet.column_dimensions["E"].width = 12.29
            active_sheet.column_dimensions["F"].width = 12.29
            active_sheet.column_dimensions["G"].width = 12.29
            
            for row in active_sheet[f"{dimensions}"]:
                for cell in row:
                    cell.style = formatted_cells
                    cell.style = formatted_cells
            
            for row in active_sheet["B2:G6"]:
                for cell in row:
                    cell.number_format = "0.000"
            
            for cell in header_row:
                cell.style = header_sec1
        
        else:
            active_sheet.column_dimensions["A"].width = 12.29
            active_sheet.column_dimensions["B"].width = 11.72
            active_sheet.column_dimensions["C"].width = 11.58
            active_sheet.column_dimensions["D"].width = 16.01
            active_sheet.column_dimensions["E"].width = 13.29
            active_sheet.column_dimensions["F"].width = 11.01
            active_sheet.column_dimensions["G"].width = 5.29
            active_sheet.column_dimensions["H"].width = 9.15
            active_sheet.column_dimensions["I"].width = 8.72
            active_sheet.column_dimensions["J"].width = 8.72
            active_sheet.column_dimensions["K"].width = 8.72
            active_sheet.column_dimensions["L"].width = 8.72
            active_sheet.column_dimensions["M"].width = 8.86
            active_sheet.column_dimensions["N"].width = 8.72
            active_sheet.column_dimensions["O"].width = 8.72
            active_sheet.column_dimensions["P"].width = 9.72
            active_sheet.column_dimensions["Q"].width = 9.72
            active_sheet.column_dimensions["R"].width = 9.72
            active_sheet.column_dimensions["S"].width = 9.72
            active_sheet.column_dimensions["T"].width = 9.72
            active_sheet.column_dimensions["U"].width = 9.72
            active_sheet.column_dimensions["V"].width = 9.72
            active_sheet.column_dimensions["W"].width = 9.72
            
            for row in active_sheet[f"{dimensions}"]:
                for cell in row:
                    cell.style = formatted_cells
                    cell.style = formatted_cells
            
            for row in active_sheet["M2:M1000"]:
                for cell in row:
                    cell.number_format = "0.00%"
            
            for row in active_sheet["N2:O1000"]:
                for cell in row:
                    cell.number_format = "0.00"
            
            for row in active_sheet["P2:T1000"]:
                for cell in row:
                    cell.number_format = "0.000"
            
            for cell in header_row:
                cell.style = header
    
    new_sheet = workbook.create_sheet("Author")
    
    author_list = ["Abhiyan Sapkota", "AJ Newcomer", "Alexa Sanchez Ochoa", "Anmol Ahmad", "Armando Gonzalez Hernandez", "Brandon Sousa", "Chris Kuwatani", "Coltyce Sanders", "David Cervantes", "David Weaver", "Devin Chang", "Jennifer Bunnell", "John Moestopo", "Kiara Davis", "Miguel Llamas", "Monika Sipa-Skorka", "Nina Villar", "Prem Dhar", "Remi Rodberg", "Richard Jankovics", "Ruben Lozano", "Samuel Yarman", "Tony Soares", "Truc Tran", "Ysidra van Kempen", "Zachary A. Coustier", "Zoe Moestopo"]
    
    new_sheet["A1"] = f"Brought to you by: {author_list[random.randrange(0, len(author_list))]}"
    new_sheet["A1"].font = Font(name = "Arial", sz = 72)
    
    workbook.save(filename = filepath)

def center_window(primary_window, new_window):
    
    
    primary_window_x, primary_window_y = primary_window.current_location()
    primary_window_width, primary_window_height = primary_window.size
    new_win_width, new_win_height = new_window.size
    
    new_x_location = (primary_window_width // 2) - (new_win_width // 2) + primary_window_x
    new_y_location = (primary_window_height // 2) - (new_win_height // 2) + primary_window_y
    
    new_window.move(new_x_location, new_y_location)
    new_window.location = (new_x_location, new_y_location)
    
    return new_x_location, new_y_location

def summary_compare(cwd):
    """
    Formats the workbook for readability.
    
    Args:
        cwd (str): String filepath to the raw workbook
    """
    
    sum_compare_out = [
        [
            sg.Text("Output Directory for Summary Table Comparison: "),
            sg.Input(key = "-SUM_COMPARE_OUT-"),
            sg.FolderBrowse(key = "-SUM_COMPARE_OUT-", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    sum_compare_input = [
        [
            sg.Text("Summary Workbook 1: "),
            sg.Input(key = "-SUM_COMPARE_IN_1-"),
            sg.FileBrowse(key = "-SC1_IN_BROWSE-", button_color = ("black", "#D3D3D3"))
        ],
        
        [
            sg.Text("Summary Workbook 2: "),
            sg.Input(key = "-SUM_COMPARE_IN_2-"),
            sg.FileBrowse(key = "-SC2_IN_BROWSE-", button_color = ("black", "#D3D3D3"))
        ],
    ]
    
    sum_compare_layout = [
        [
            sg.Column(sum_compare_out, element_justification = "right"),
            sg.VerticalSeparator(),
            sg.Column(sum_compare_input, element_justification = "left"),
        ],
        
        [
            sg.Exit(button_color = ("white", "red")),
            sg.Button("Start Comparison", key = "-COMPARE_NOW-", button_color = ("black", "#D3D3D3")),
        ]
    ]
    
    sum_compare_window = sg.Window("Summary Comparison", sum_compare_layout)
    
    while True:
        sc_event, sc_values = sum_compare_window.read()
        
        if sc_event in (sg.WINDOW_CLOSED, "Exit"):
            break
        
        if sc_values["-SUM_COMPARE_OUT-"] != "":
            sum_compare_workbook_filename = parse_summary_comparison_workbook_name(sc_values["-SUM_COMPARE_IN_1-"])
            sum_compare_workbook_directory = sc_values["-SUM_COMPARE_OUT-"]
            sum_compare_workbook_filepath = os.path.join(sum_compare_workbook_directory, sum_compare_workbook_filename) # sum_compare_workbook_directory + "/" + sum_compare_workbook_filename
            
            sum_com_wb1 = sc_values["-SUM_COMPARE_IN_1-"].split("/")
            sum_com_wb2 = sc_values["-SUM_COMPARE_IN_2-"].split("/")
            
            sum_com_wb1_name = sum_com_wb1[-1]
            sum_com_wb2_name = sum_com_wb2[-1]
            
        if sc_event == "-COMPARE_NOW-":
            stuff_to_compare = Compare(sum_workbook_1 = sc_values["-SUM_COMPARE_IN_1-"], sum_workbook_2 = sc_values["-SUM_COMPARE_IN_2-"] , comparison_filepath = sum_compare_workbook_filepath, log_dir = cwd)
            
            stuff_to_compare.compare_workbooks()
            stuff_to_compare.compare_format_results()
            
            summary_compare_excel(sum_compare_workbook_filepath, sum_com_wb1_name, sum_com_wb2_name)
            
            sg.popup("Done", button_color = ("black", "#D3D3D3"))
            
            sum_compare_window.close()
    
    sum_compare_window.close()

def gsm_tech_window(primary_window):
    """
    Window for technology selection

    Returns:
        list: List of selected technologies
    """
    
    gsm_tech_list = [
        [sg.Checkbox("GSM 850", key = "GSM 850")],
        [sg.Checkbox("GSM E-900", key = "GSM E-900")],
        [sg.Checkbox("GSM R-900", key = "GSM R-900")],
        [sg.Checkbox("GSM 1800", key = "GSM 1800")],
        [sg.Checkbox("GSM 1900", key = "GSM 1900")],
        
        [
            sg.Submit(button_color = ("black", "#D3D3D3")),
            sg.Exit(button_text = "Close", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    gsm_window = sg.Window("Select all that apply:", gsm_tech_list, finalize = True)
    center_window(primary_window, new_window = gsm_window)
    
    while True:
        gsm_event, gsm_values = gsm_window.read()
        # print(gsm_event, gsm_values)
        if gsm_event in (sg.WIN_CLOSED, "Exit"):
            break
        
        if gsm_event == "Submit":
            checked_gsm = gsm_values
            
            new_gsm_dict = {key:value for key, value in checked_gsm.items() if value != False}
            
            final_gsm_list = list(new_gsm_dict.keys())
            
            gsm_window.close()
    
    gsm_window.close()
    
    return final_gsm_list

def wcdma_tech_window(primary_window):
    """
    Window for technology selection

    Returns:
        list: List of selected technologies
    """
    
    wcdma_tech_list = [
        [sg.Checkbox("W-CDMA B1", key = "W-CDMA B1")],
        [sg.Checkbox("W-CDMA B2", key = "W-CDMA B2")],
        [sg.Checkbox("W-CDMA B4", key = "W-CDMA B4")],
        [sg.Checkbox("W-CDMA B5", key = "W-CDMA B5")],
        [sg.Checkbox("W-CDMA B8", key = "W-CDMA B8")],
        
        [
            sg.Submit(button_color = ("black", "#D3D3D3")),
            sg.Exit(button_text = "Close", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    wcdma_window = sg.Window("Select all that apply:", wcdma_tech_list, finalize = True)
    center_window(primary_window, new_window = wcdma_window)
    
    while True:
        wcdma_event, wcdma_values = wcdma_window.read()
        # print(wcdma_event, wcdma_values)
        if wcdma_event in (sg.WIN_CLOSED, "Exit"):
            break
        
        if wcdma_event == "Submit":
            checked_wcdma = wcdma_values
            
            new_wcdma_dict = {key:value for key, value in checked_wcdma.items() if value != False}
            
            final_wcdma_list = list(new_wcdma_dict.keys())
            
            wcdma_window.close()
    
    wcdma_window.close()
    
    return final_wcdma_list

def lte_tech_window(primary_window):
    """
    Window for technology selection

    Returns:
        list: List of selected technologies
    """
    
    lte_tech_1 = [
        [sg.Checkbox("LTE B1", key = "LTE B1")],
        [sg.Checkbox("LTE B2", key = "LTE B2")],
        [sg.Checkbox("LTE B3", key = "LTE B3")],
        [sg.Checkbox("LTE B4", key = "LTE B4")],
        [sg.Checkbox("LTE B5", key = "LTE B5")],
        [sg.Checkbox("LTE B7", key = "LTE B7")],
        [sg.Checkbox("LTE B8", key = "LTE B8")],
        [sg.Checkbox("LTE B11", key = "LTE B11")],
        [sg.Checkbox("LTE B12", key = "LTE B12")],
        [sg.Checkbox("LTE B13", key = "LTE B13")],
    ]
    
    lte_tech_2 = [
        [sg.Checkbox("LTE B14", key = "LTE B14")],
        [sg.Checkbox("LTE B17", key = "LTE B17")],
        [sg.Checkbox("LTE B18", key = "LTE B18")],
        [sg.Checkbox("LTE B19", key = "LTE B19")],
        [sg.Checkbox("LTE B20", key = "LTE B20")],
        [sg.Checkbox("LTE B21", key = "LTE B21")],
        [sg.Checkbox("LTE B22", key = "LTE B22")],
        [sg.Checkbox("LTE B24", key = "LTE B24")],
        [sg.Checkbox("LTE B25", key = "LTE B25")],
        [sg.Checkbox("LTE B26", key = "LTE B26")],
    ]
    
    lte_tech_3 = [
        [sg.Checkbox("LTE B27", key = "LTE B27")],
        [sg.Checkbox("LTE B28", key = "LTE B28")],
        [sg.Checkbox("LTE B30", key = "LTE B30")],
        [sg.Checkbox("LTE B31", key = "LTE B31")],
        [sg.Checkbox("LTE B33", key = "LTE B33")],
        [sg.Checkbox("LTE B34", key = "LTE B34")],
        [sg.Checkbox("LTE B35", key = "LTE B35")],
        [sg.Checkbox("LTE B36", key = "LTE B36")],
        [sg.Checkbox("LTE B37", key = "LTE B37")],
        [sg.Checkbox("LTE B38", key = "LTE B38")],
    ]
    
    lte_tech_4 = [
        [sg.Checkbox("LTE B39", key = "LTE B39")],
        [sg.Checkbox("LTE B40 (Block A)", key = "LTE B40 (Block A)")],
        [sg.Checkbox("LTE B40 (Block B)", key = "LTE B40 (Block B)")],
        [sg.Checkbox("LTE B40", key = "LTE B40")],
        [sg.Checkbox("LTE B41 FCC PC3", key = "LTE B41 FCC PC3")],
        [sg.Checkbox("LTE B41 FCC PC2", key = "LTE B41 FCC PC2")],
        [sg.Checkbox("LTE B42", key = "LTE B42")],
        [sg.Checkbox("LTE B43", key = "LTE B43")],
        [sg.Checkbox("LTE B44", key = "LTE B44")],
        [sg.Checkbox("LTE B45", key = "LTE B45")],
    ]
    
    lte_tech_5 = [
        [sg.Checkbox("LTE B46 NAR", key = "LTE B46 NAR")],
        [sg.Checkbox("LTE B46 CE", key = "LTE B46 CE")],
        [sg.Checkbox("LTE B47", key = "LTE B47")],
        [sg.Checkbox("LTE B48", key = "LTE B48")],
        [sg.Checkbox("LTE B49", key = "LTE B49")],
        [sg.Checkbox("LTE B50", key = "LTE B50")],
        [sg.Checkbox("LTE B51", key = "LTE B51")],
        [sg.Checkbox("LTE B52", key = "LTE B52")],
        [sg.Checkbox("LTE B53", key = "LTE B53")],
        [sg.Checkbox("LTE B65", key = "LTE B65")],
    ]
    
    lte_tech_6 = [
        [sg.Checkbox("LTE B66", key = "LTE B66")],
        [sg.Checkbox("LTE B68", key = "LTE B68")],
        [sg.Checkbox("LTE B70", key = "LTE B70")],
        [sg.Checkbox("LTE B71", key = "LTE B71")],
        [sg.Checkbox("LTE B72", key = "LTE B72")],
        [sg.Checkbox("LTE B73", key = "LTE B73")],
        [sg.Checkbox("LTE B74", key = "LTE B74")],
        [sg.Checkbox("LTE B85", key = "LTE B85")],
        [sg.Checkbox("LTE B87", key = "LTE B87")],
        [sg.Checkbox("LTE B88", key = "LTE B88")],
    ]
    
    lte_layout =[
        [
            sg.Column(lte_tech_1, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(lte_tech_2, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(lte_tech_3, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(lte_tech_4, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(lte_tech_5, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(lte_tech_6, element_justification = "left"),
        ],
        
        [
            sg.Submit(button_color = ("black", "#D3D3D3")),
            sg.Exit(button_text = "Close", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    lte_window = sg.Window("Select all that apply:", lte_layout, finalize = True)
    center_window(primary_window, new_window = lte_window)
    
    while True:
        lte_event, lte_values = lte_window.read()
        # print(lte_event, lte_values)
        if lte_event in (sg.WIN_CLOSED, "Exit"):
            break
        
        if lte_event == "Submit":
            checked_lte = lte_values
            
            new_lte_dict = {key:value for key, value in checked_lte.items() if value != False}
            
            final_lte_list = list(new_lte_dict.keys())
            
            lte_window.close()
    
    lte_window.close()
    
    return final_lte_list

def fr1_tech_window(primary_window):
    """
    Window for technology selection

    Returns:
        list: List of selected technologies
    """
    
    fr1_tech_1 = [
        [sg.Checkbox("FR1 n1", key = "FR1 n1")],
        [sg.Checkbox("FR1 n2", key = "FR1 n2")],
        [sg.Checkbox("FR1 n3", key = "FR1 n3")],
        [sg.Checkbox("FR1 n5", key = "FR1 n5")],
        [sg.Checkbox("FR1 n7", key = "FR1 n7")],
        [sg.Checkbox("FR1 n8", key = "FR1 n8")],
        [sg.Checkbox("FR1 n12", key = "FR1 n12")],
        [sg.Checkbox("FR1 n13", key = "FR1 n13")],
        [sg.Checkbox("FR1 n14", key = "FR1 n14")],
        [sg.Checkbox("FR1 n18", key = "FR1 n18")],
        [sg.Checkbox("FR1 n20", key = "FR1 n20")],
        [sg.Checkbox("FR1 n24", key = "FR1 n24")],
        [sg.Checkbox("FR1 n25", key = "FR1 n25")],
        [sg.Checkbox("FR1 n26", key = "FR1 n26")],
    ]
    
    fr1_tech_2 = [
        [sg.Checkbox("FR1 n28", key = "FR1 n28")],
        [sg.Checkbox("FR1 n30", key = "FR1 n30")],
        [sg.Checkbox("FR1 n34", key = "FR1 n34")],
        [sg.Checkbox("FR1 n38", key = "FR1 n38")],
        [sg.Checkbox("FR1 n40", key = "FR1 n40")],
        [sg.Checkbox("FR1 n40 (Block A)", key = "FR1 n40 (Block A)")],
        [sg.Checkbox("FR1 n40 (Block B)", key = "FR1 n40 (Block B)")],
        [sg.Checkbox("FR1 n40 CE", key = "FR1 n40 CE")],
        [sg.Checkbox("FR1 n41 PC3", key = "FR1 n41 PC3")],
        [sg.Checkbox("FR1 n41 PC2", key = "FR1 n41 PC2")],
        [sg.Checkbox("FR1 n41 CE PC3", key = "FR1 n41 CE PC3")],
        [sg.Checkbox("FR1 n41 CE PC2", key = "FR1 n41 CE PC2")],
        [sg.Checkbox("FR1 n46", key = "FR1 n46")],
        [sg.Checkbox("FR1 n46 CE", key = "FR1 n46 CE")],
    ]
    
    fr1_tech_3 = [
        [sg.Checkbox("FR1 n47", key = "FR1 n47")],
        [sg.Checkbox("FR1 n48", key = "FR1 n48")],
        [sg.Checkbox("FR1 n48 CE", key = "FR1 n48 CE")],
        [sg.Checkbox("FR1 n50", key = "FR1 n50")],
        [sg.Checkbox("FR1 n51", key = "FR1 n51")],
        [sg.Checkbox("FR1 n53", key = "FR1 n53")],
        [sg.Checkbox("FR1 n65", key = "FR1 n65")],
        [sg.Checkbox("FR1 n66", key = "FR1 n66")],
        [sg.Checkbox("FR1 n70", key = "FR1 n70")],
        [sg.Checkbox("FR1 n71", key = "FR1 n71")],
        [sg.Checkbox("FR1 n74", key = "FR1 n74")],
        [sg.Checkbox("FR1 n77 PC3", key = "FR1 n77 PC3")],
        [sg.Checkbox("FR1 n77 CE PC3", key = "FR1 n77 CE PC3")],
        [sg.Checkbox("FR1 n77 (Block A) PC3", key = "FR1 n77 (Block A) PC3")],
    ]
    
    fr1_tech_4 = [
        [sg.Checkbox("FR1 n77 (Block B) PC3", key = "FR1 n77 (Block B) PC3")],
        [sg.Checkbox("FR1 n77 CE (Block B) PC3", key = "FR1 n77 CE (Block B) PC3")],
        [sg.Checkbox("FR1 n77 (Block C) PC3", key = "FR1 n77 (Block C) PC3")],
        [sg.Checkbox("FR1 n77 CE (Block C) PC3", key = "FR1 n77 CE (Block C) PC3")],
        [sg.Checkbox("FR1 n77 PC2", key = "FR1 n77 PC2")],
        [sg.Checkbox("FR1 n77 CE PC2", key = "FR1 n77 CE PC2")],
        [sg.Checkbox("FR1 n77 (Block A) PC2", key = "FR1 n77 (Block A) PC2")],
        [sg.Checkbox("FR1 n77 (Block B) PC2", key = "FR1 n77 (Block B) PC2")],
        [sg.Checkbox("FR1 n77 CE (Block B) PC2", key = "FR1 n77 CE (Block B) PC2")],
        [sg.Checkbox("FR1 n77 (Block C) PC2", key = "FR1 n77 (Block C) PC2")],
        [sg.Checkbox("FR1 n77 CE (Block C) PC2", key = "FR1 n77 CE (Block C) PC2")],
        [sg.Checkbox("FR1 n78 PC3", key = "FR1 n78 PC3")],
        [sg.Checkbox("FR1 n78 (Block A) PC3", key = "FR1 n78 (Block A) PC3")],
        [sg.Checkbox("FR1 n78 (Block B) PC3", key = "FR1 n78 (Block B) PC3")],
    ]
    
    fr1_tech_5 = [
        [sg.Checkbox("FR1 n78 (Block C) PC3", key = "FR1 n78 (Block C) PC3")],
        [sg.Checkbox("FR1 n78 IC PC3", key = "FR1 n78 IC PC3")],
        [sg.Checkbox("FR1 n78 CE PC3", key = "FR1 n78 CE PC3")],
        [sg.Checkbox("FR1 n78 PC2", key = "FR1 n78 PC2")],
        [sg.Checkbox("FR1 n78 (Block A) PC2", key = "FR1 n78 (Block A) PC2")],
        [sg.Checkbox("FR1 n78 (Block B) PC2", key = "FR1 n78 (Block B) PC2")],
        [sg.Checkbox("FR1 n78 (Block C) PC2", key = "FR1 n78 (Block C) PC2")],
        [sg.Checkbox("FR1 n78 IC PC2", key = "FR1 n78 IC PC2")],
        [sg.Checkbox("FR1 n78 CE PC2", key = "FR1 n78 CE PC2")],
        [sg.Checkbox("FR1 n79", key = "FR1 n79")],
        [sg.Checkbox("FR1 n79 CE", key = "FR1 n79 CE")],
        [sg.Checkbox("FR1 n79 (Narrow) PC3", key = "FR1 n79 (Narrow) PC3")],
        [sg.Checkbox("FR1 n79 (Narrow) PC2", key = "FR1 n79 (Narrow) PC2")],
        [sg.Checkbox("FR1 n85", key = "FR1 n85")],
        [sg.Checkbox("FR1 n90", key = "FR1 n90")],
    ]
    
    fr1_tech_6 = [
        [sg.Checkbox("FR1 n90 CE", key = "FR1 n90 CE")],
        [sg.Checkbox("FR1 n91", key = "FR1 n91")],
        [sg.Checkbox("FR1 n92", key = "FR1 n92")],
        [sg.Checkbox("FR1 n93", key = "FR1 n93")],
        [sg.Checkbox("FR1 n94", key = "FR1 n94")],
        [sg.Checkbox("FR1 n96", key = "FR1 n96")],
        [sg.Checkbox("FR1 n96 CE", key = "FR1 n96 CE")],
        [sg.Checkbox("FR1 n100", key = "FR1 n100")],
        [sg.Checkbox("FR1 n101", key = "FR1 n101")],
        [sg.Checkbox("FR1 n102", key = "FR1 n102")],
        [sg.Checkbox("FR1 n102 CE", key = "FR1 n102 CE")],
        [sg.Checkbox("FR1 n104", key = "FR1 n104")],
        [sg.Checkbox("FR1 n104 CE", key = "FR1 n104 CE")],
        [sg.Checkbox("FR1 n255", key = "FR1 n255")],
        [sg.Checkbox("FR1 n256", key = "FR1 n256")],
    ]
    
    fr1_layout =[
        [
            sg.Column(fr1_tech_1, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(fr1_tech_2, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(fr1_tech_3, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(fr1_tech_4, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(fr1_tech_5, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(fr1_tech_6, element_justification = "left"),
        ],
        
        [
            sg.Submit(button_color = ("black", "#D3D3D3")),
            sg.Exit(button_text = "Close", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    fr1_window = sg.Window("Select all that apply:", fr1_layout, finalize = True)
    center_window(primary_window, new_window = fr1_window)
    
    while True:
        fr1_event, fr1_values = fr1_window.read()
        # print(fr1_event, fr1_values)
        if fr1_event in (sg.WIN_CLOSED, "Exit"):
            break
        
        if fr1_event == "Submit":
            checked_fr1 = fr1_values
            
            new_fr1_dict = {key:value for key, value in checked_fr1.items() if value != False}
            
            final_fr1_list = list(new_fr1_dict.keys())
            
            fr1_window.close()
    
    fr1_window.close()
    
    return final_fr1_list

def wlan_tech_window(primary_window):
    """
    Window for technology selection

    Returns:
        list: List of selected technologies
    """
    
    wlan_tech_list = [
        [sg.Checkbox("Wi-Fi 2.4 GHz", key = "Wi-Fi 2.4 GHz")],
        [sg.Checkbox("Wi-Fi 5.2 GHz", key = "Wi-Fi 5.2 GHz")],
        [sg.Checkbox("Wi-Fi 5.3 GHz", key = "Wi-Fi 5.3 GHz")],
        [sg.Checkbox("Wi-Fi 5.5 GHz", key = "Wi-Fi 5.5 GHz")],
        [sg.Checkbox("Wi-Fi 5.8 GHz", key = "Wi-Fi 5.8 GHz")],
        [sg.Checkbox("Wi-Fi 5.9 GHz", key = "Wi-Fi 5.9 GHz")],
    ]
    
    wlan_6e_tech_list = [
        [sg.Checkbox("Wi-Fi 6E (Aggregated)", key = "Wi-Fi 6E")],
        [sg.Checkbox("U-NII 5", key = "U-NII 5")],
        [sg.Checkbox("U-NII 6", key = "U-NII 6")],
        [sg.Checkbox("U-NII 7", key = "U-NII 7")],
        [sg.Checkbox("U-NII 8", key = "U-NII 8")],
    ]
    
    wlan_tech_layout = [
        [
            sg.Column(wlan_tech_list, element_justification = "left"),
            sg.VerticalSeparator(),
            sg.Column(wlan_6e_tech_list, element_justification = "left"),
        ],
        
        [
            sg.Submit(button_color = ("black", "#D3D3D3")),
            sg.Exit(button_text = "Close", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    wlan_window = sg.Window("Select all that apply:", wlan_tech_layout, finalize = True)
    center_window(primary_window, new_window = wlan_window)
    
    while True:
        wlan_event, wlan_values = wlan_window.read()
        # print(wlan_event, wlan_values)
        if wlan_event in (sg.WIN_CLOSED, "Exit"):
            break
        
        if wlan_event == "Submit":
            checked_wlan = wlan_values
            
            new_wlan_dict = {key:value for key, value in checked_wlan.items() if value != False}
            
            final_wlan_list = list(new_wlan_dict.keys())
            
            wlan_window.close()
    
    wlan_window.close()
    
    return final_wlan_list

def bt_tech_window(primary_window):
    """
    Window for technology selection

    Returns:
        list: List of selected technologies
    """
    
    bluetooth_tech_list = [
        [sg.Checkbox("Bluetooth (2.4 GHz)", key = "Bluetooth (2.4 GHz)")],
        [sg.Checkbox("Bluetooth (NB U-NII 1)", key = "Bluetooth (NB U-NII 1)")],
        [sg.Checkbox("Bluetooth (NB U-NII 3)", key = "Bluetooth (NB U-NII 3)")],
        
        [
            sg.Submit(button_color = ("black", "#D3D3D3")),
            sg.Exit(button_text = "Close", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    bluetooth_window = sg.Window("Select all that apply:", bluetooth_tech_list, finalize = True)
    center_window(primary_window, new_window = bluetooth_window)
    
    while True:
        bluetooth_event, bluetooth_values = bluetooth_window.read()
        # print(bluetooth_event, bluetooth_values)
        if bluetooth_event in (sg.WIN_CLOSED, "Exit"):
            break
        
        if bluetooth_event == "Submit":
            checked_bluetooth = bluetooth_values
            
            new_bluetooth_dict = {key:value for key, value in checked_bluetooth.items() if value != False}
            
            final_bluetooth_list = list(new_bluetooth_dict.keys())
            
            bluetooth_window.close()
    
    bluetooth_window.close()
    
    return final_bluetooth_list

def thread_tech_window(primary_window):
    """
    Window for technology selection

    Returns:
        list: List of selected technologies
    """
    
    thread_tech_list = [
        [sg.Checkbox("802.15.4", key = "802.15.4")],
        [sg.Checkbox("802.15.4ab", key = "802.15.4ab")],
        
        [
            sg.Submit(button_color = ("black", "#D3D3D3")),
            sg.Exit(button_text = "Close", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    thread_window = sg.Window("Select all that apply:", thread_tech_list, finalize = True)
    center_window(primary_window, new_window = thread_window)
    
    while True:
        thread_event, thread_values = thread_window.read()
        # print(thread_event, thread_values)
        if thread_event in (sg.WIN_CLOSED, "Exit"):
            break
        
        if thread_event == "Submit":
            checked_thread = thread_values
            
            new_thread_dict = {key:value for key, value in checked_thread.items() if value != False}
            
            final_thread_list = list(new_thread_dict.keys())
            
            thread_window.close()
    
    thread_window.close()
    
    return final_thread_list

def mss_tech_window(primary_window):
    """
    Window for technology selection

    Returns:
        list: List of selected technologies
    """
    
    mss_tech_list = [
        [
            sg.Checkbox("MSS (L-Band)", key = "MSS (L-Band)"),
            sg.Checkbox("NTN L-Band", key = "NTN L-Band"),
            sg.Checkbox("NTN S-Band", key = "NTN S-Band")
        ],
        
        [
            sg.Submit(button_color = ("black", "#D3D3D3")),
            sg.Exit(button_text = "Close", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    mss_window = sg.Window("Select all that apply:", mss_tech_list, finalize = True)
    center_window(primary_window, new_window = mss_window)
    
    while True:
        mss_event, mss_values = mss_window.read()
        # print(mss_event, mss_values)
        if mss_event in (sg.WIN_CLOSED, "Exit"):
            break
        
        if mss_event == "Submit":
            checked_mss = mss_values
            
            new_mss_dict = {key:value for key, value in checked_mss.items() if value != False}
            
            final_mss_list = list(new_mss_dict.keys())
            
            mss_window.close()
    
    mss_window.close()
    
    return final_mss_list

def reported_results(primary_window, rr_out, nwtl, wtl, data, rwf, swf, stxwf, cwd):
    """
    Collects all the variables needed to create the reported results raw
    workbook.
    
    Args:
        rr_out (str): String filepath to the workbook output filepath
        nwtl (list): List of non-Wi-Fi technologies
        wtl (list): List of Wi-Fi technologies
        data (dataframe): Data collected from SAR workbooks
        rwf (str): String filepath for the reported results workbook
        swf (str): String filepath for the summary results workbook
        stxwf (str): String filepath for the TAS workbook
        cwd (str): String filepath of the current working directory
    """
    
    yes_no = [
        [sg.Text("Has all the pertinent information been entered?")],
        
        [
            sg.Button("Yes", key = "Yes", button_color = ("black", "#D3D3D3")),
            sg.Button("AJ's Button", key = "AJ", button_color = ("pink", "#3CFF00"), font = ("Helvetica", 11, "bold")),
            sg.Exit(button_text = "No", button_color = ("white", "red"))
        ]
    ]
    
    reported_results = sg.Window("Generate Reported Restults Workbook", yes_no, element_justification = "center", finalize = True)
    center_window(primary_window, new_window = reported_results)
    
    if rr_out != "":
        while True:
            rr_event, rr_values = reported_results.read()
            # print(rr_event, rr_values)
            if rr_event in (sg.WIN_CLOSED, "Exit"):
                break
            
            if rr_event == "Yes":
                if not nwtl:
                    print("There is no non-Wi-Fi Data.")
                else:
                    stuff_for_not_wlan = Not_WiFi(data = data, tech_list = nwtl, transmitter_names = "", exposure_conditions = "", reported_results_filepath = rwf, summary_results_filepath = swf, smtx_results_filepath = stxwf, log_dir = cwd)
                    
                    stuff_for_not_wlan.reported_tech_results()
                
                if not wtl:
                    print("There is no Wi-Fi Data.")
                else:
                    stuff_for_wlan = Wlan(data = data, tech_list = wtl, transmitter_names = "", exposure_conditions = "", reported_results_filepath = rwf, summary_results_filepath = swf, smtx_results_filepath = stxwf, log_dir = cwd)
                    
                    stuff_for_wlan.reported_tech_results()
                
                reported_excel(rwf)
                
                sg.popup("Done", button_color = ("black", "#D3D3D3"), location = (center_window(primary_window, new_window = reported_results)))
                
                reported_results.close()
                
            if rr_event == "AJ":
                not_wlan = ["GSM 850", "GSM 1900", "W-CDMA B2", "W-CDMA B4", "W-CDMA B5", "LTE B2", "LTE B4", "LTE B5", "LTE B7", "LTE B12", "LTE B13", "LTE B14", "LTE B25", "LTE B26", "LTE B30", "LTE B41 PC3", "LTE B48", "LTE B53", "LTE B66", "LTE B71", "FR1 n5", "FR1 n7", "FR1 n12", "FR1 n14", "FR1 n25", "FR1 n26", "FR1 n30", "FR1 n41 PC3", "FR1 n48", "FR1 n53", "FR1 n66", "FR1 n70", "FR1 n71", "FR1 n77 (Block A) PC3", "FR1 n77 (Block C) PC3", "FR1 n78 IC PC3", "FR1 n79 (Narrow) PC3", "Bluetooth (2.4 GHz)", "Bluetooth (NB U-NII 1)", "Bluetooth (NB U-NII 3)", "802.15.4", "802.15.4ab", "MSS (L-Band)"]
                
                wlan = ["Wi-Fi 2.4 GHz", "Wi-Fi 5.2 GHz", "Wi-Fi 5.3 GHz", "Wi-Fi 5.5 GHz", "Wi-Fi 5.8 GHz", "U-NII 5", "U-NII 6", "U-NII 7", "U-NII 8"]
                
                stuff_for_not_wlan = Not_WiFi(data = data, tech_list = not_wlan, transmitter_names = "", exposure_conditions = "", reported_results_filepath = rwf, summary_results_filepath = swf, smtx_results_filepath = stxwf, log_dir = cwd)
                stuff_for_wlan = Wlan(data = data, tech_list = wlan, transmitter_names = "", exposure_conditions = "", reported_results_filepath = rwf, summary_results_filepath = swf, smtx_results_filepath = stxwf, log_dir = cwd)
                
                stuff_for_not_wlan.reported_tech_results()
                stuff_for_wlan.reported_tech_results()
                
                reported_excel(rwf)
                
                sg.popup("Done", button_color = ("black", "#D3D3D3"), location = (center_window(primary_window, new_window = reported_results)))
                
                reported_results.close()
                
        reported_results.close()
    
    else:
        sg.popup("The Reported Results workbook directory has not been entered;\nplease enter the appropriate directory", button_color = ("black", "#D3D3D3"), location = (center_window(primary_window, new_window = reported_results)))

def summary_results(primary_window, sr_out, nwtl, wtl, data, rwf, swf, stxwf, cwd, exl):
    """
    Collects all the variables needed to create the reported results raw
    workbook.
    
    Args:
        rr_out (str): String filepath to the workbook output filepath
        nwtl (list): List of non-Wi-Fi technologies
        wtl (list): List of Wi-Fi technologies
        data (dataframe): Data collected from SAR workbooks
        rwf (str): String filepath for the reported results workbook
        swf (str): String filepath for the summary results workbook
        stxwf (str): String filepath for the TAS workbook
        cwd (str): String filepath of the current working directory
        exl (list): List of applicable exposure conditions
    """
    
    transmitter_names = [
        [
            sg.Text("Enter the naming convention per transmitter.")
        ],
        
        [
            sg.Text("Transmitter 1:"),
            sg.Input(key = "-TRANS_1-")
        ],
        
        [
            sg.Text("Transmitter 2:"),
            sg.Input(key = "-TRANS_2-")
        ],
        
        [
            sg.Text("Transmitter 3:"),
            sg.Input(key = "-TRANS_3-")
        ],
        
        [
            sg.Text("Transmitter 4:"),
            sg.Input(key = "-TRANS_4-")
        ],
        
        [
            sg.Text("Transmitter 5:"),
            sg.Input(key = "-TRANS_5-")
        ],
        
        [
            sg.Text("Transmitter 6:"),
            sg.Input(key = "-TRANS_6-")
        ],
        
        [
            sg.Text("Transmitter 7:"),
            sg.Input(key = "-TRANS_7-")
        ],
        
        [
            sg.Text("Transmitter 8:"),
            sg.Input(key = "-TRANS_8-")
        ],
        
        [
            sg.Text("Transmitter 9:"),
            sg.Input(key = "-TRANS_9-")
        ],
        
        [
            sg.Submit(button_color = ("black", "#D3D3D3")),
            sg.Button("AJ's Button", key = "AJ", button_color = ("pink", "#3CFF00"), font = ("Helvetica", 11, "bold")),
            sg.Exit(button_text = "Close", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    summary_table = sg.Window("Generate Summary Table", transmitter_names, element_justification = "left", finalize = True)
    center_window(primary_window, new_window = summary_table)
    
    if sr_out != "":
        while True:
            sr_event, sr_values = summary_table.read()
            # print(sr_event, sr_values)
            if sr_event in (sg.WIN_CLOSED, "Exit"):
                break
            
            if sr_event == "Submit":
                # print(sr_values)
                filled_trans = sr_values
                
                new_trans_dict = {key:value for key, value in filled_trans.items() if value != ""}
                
                final_trans_list = list(new_trans_dict.values())
                
                if not nwtl:
                    print("There is no non-Wi-Fi Data.")
                else:
                    stuff_for_not_wlan = Not_WiFi(data = data, tech_list = nwtl, transmitter_names = final_trans_list, exposure_conditions = exl, reported_results_filepath = rwf, summary_results_filepath = swf, smtx_results_filepath = stxwf, log_dir = cwd)
                    
                    stuff_for_not_wlan.summary_tech_results()
                
                if not wtl:
                    print("There is no Wi-Fi Data.")
                else:
                    stuff_for_wlan = Wlan(data = data, tech_list = wtl, transmitter_names = final_trans_list, exposure_conditions = exl, reported_results_filepath = rwf, summary_results_filepath = swf, smtx_results_filepath = stxwf, log_dir = cwd)
                    
                    stuff_for_wlan.summary_tech_results()
                
                summary_excel(swf)
                
                sg.popup("Done", button_color = ("black", "#D3D3D3"), location = (center_window(primary_window, new_window = summary_table)))
                
                summary_table.close()
                
            if sr_event == "AJ":
                not_wlan = ["GSM 850", "GSM 1900", "W-CDMA B2", "W-CDMA B4", "W-CDMA B5", "LTE B2", "LTE B4", "LTE B5", "LTE B7", "LTE B12", "LTE B13", "LTE B14", "LTE B25", "LTE B26", "LTE B30", "LTE B41 FCC PC3", "LTE B41 IC PC3", "LTE B48", "LTE B53", "LTE B66", "LTE B71", "FR1 n5", "FR1 n7", "FR1 n12", "FR1 n14", "FR1 n25", "FR1 n26", "FR1 n30", "FR1 n41 PC3", "FR1 n48", "FR1 n53", "FR1 n66", "FR1 n70", "FR1 n71", "FR1 n77 (Block A) PC3", "FR1 n77 (Block C) PC3", "FR1 n78 IC PC3", "FR1 n79 (Narrow) PC3", "Bluetooth (2.4 GHz)", "Bluetooth (NB U-NII 1)", "Bluetooth (NB U-NII 3)", "802.15.4", "802.15.4ab", "MSS (L-Band)"]
                
                wlan = ["Wi-Fi 2.4 GHz", "Wi-Fi 5.2 GHz", "Wi-Fi 5.3 GHz", "Wi-Fi 5.5 GHz", "Wi-Fi 5.8 GHz", "U-NII 5", "U-NII 6", "U-NII 7", "U-NII 8"]
                
                filled_trans = sr_values
                
                new_trans_dict = {key:value for key, value in filled_trans.items() if value != ""}
                
                final_trans_list = list(new_trans_dict.values())
                
                stuff_for_not_wlan = Not_WiFi(data = data, tech_list = not_wlan, transmitter_names = final_trans_list, exposure_conditions = exl, reported_results_filepath = rwf, summary_results_filepath = swf, smtx_results_filepath = stxwf, log_dir = cwd)
                stuff_for_wlan = Wlan(data = data, tech_list = wlan, transmitter_names = final_trans_list, exposure_conditions = exl, reported_results_filepath = rwf, summary_results_filepath = swf, smtx_results_filepath = stxwf, log_dir = cwd)
                
                stuff_for_not_wlan.summary_tech_results()
                stuff_for_wlan.summary_tech_results()
                
                summary_excel(swf)
                
                sg.popup("Done", button_color = ("black", "#D3D3D3"), location = (center_window(primary_window, new_window = summary_table)))
                
                summary_table.close()
                
        summary_table.close()
    
    else:
        sg.popup("The Summary workbook directory has not been entered;\nplease enter the appropriate directory", button_color = ("black", "#D3D3D3"), location = (center_window(primary_window, new_window = summary_table)))

def sec1_results(primary_window, sec_out, data, fcc, ised, cwd):
    """
    Parses through the data to generate the Section 1 table and
    variant test cases.
    
    Args:
        sec_out (str): Output direcotry for the Section 1 workbook
        data (dataframe): Raw data collected from the SAR workbook(s)
        fcc (str): String filepath for the FCC workbook
        ised (str): String filepath for the ISED workbook
        cwd (str): String filepath of the current working directory
    """
    
    yes_no = [
        [sg.Text("Ready for 'OH WOW!'?")],
        
        [
            sg.Button("Yes", key = "Yes", button_color = ("black", "#D3D3D3")),
            sg.Exit(button_text = "No", button_color = ("white", "red"))
        ]
    ]
    
    sec1_results = sg.Window("Generate Section 1 Workbook", yes_no, element_justification = "center", finalize = True)
    center_window(primary_window, new_window = sec1_results)
    
    if sec_out != "":
        while True:
            s1_event, s1_values = sec1_results.read()
            # print(s1_event, s1_values)
            if s1_event in (sg.WIN_CLOSED, "Exit"):
                break
            
            if s1_event == "Yes":
                stuff_for_s1 = Sec_1(data = data, transmitter_names = "", sec_1_fcc_filepath = fcc, sec_1_ised_filepath = ised, log_dir = cwd)
                
                stuff_for_s1.section_1_fcc()
                stuff_for_s1.section_1_ised()
                
                sec1_excel(fcc)
                sec1_excel(ised)
                
                sg.popup("Done", button_color = ("black", "#D3D3D3"), location = (center_window(primary_window, new_window = sec1_results)))
                
                sec1_results.close()
        
        sec1_results.close()
    else:
        sg.popup("The Section 1 workbook directory has not been entered;\nplease enter the appropriate directory", button_color = ("black", "#D3D3D3"), location = (center_window(primary_window, new_window = sec1_results)))

def smtx_results(primary_window, stx_out, data, nwtl, wtl, rwf, swf, stxwf, cwd):
    """
    Generates a workbook with all the SAR data in one sheet
    
    Args:
        stx_out (str): Output directory for TAS workbook
        data (dataframe): Raw data from all the SAR workbooks
        nwtl (list): List of technologies that are not Wi-Fi
        wtl (list): List of technologies that are Wi-Fi
        rwf (str): String filepath for the reported resutls workbook
        swf (str): String filepath for the summary results workbook
        stxwf (str): String filepath for the TAS workbook
        cwd (str): String filepath for the current working directory
    """
    
    yes_no = [
        [sg.Text("Has all the pertinent information been entered?")],
        
        [
            sg.Button("Yes", key = "Yes", button_color = ("black", "#D3D3D3")),
            sg.Exit(button_text = "No", button_color = ("white", "red"))
        ]
    ]
    
    smtx_results = sg.Window("Generate Smart Transmit Restults Workbook", yes_no, element_justification = "center", finalize = True)
    center_window(primary_window, new_window = smtx_results)
    
    if stx_out != "":
        while True:
            smtx_event, smtx_values = smtx_results.read()
            # print(rr_event, rr_values)
            if smtx_event in (sg.WIN_CLOSED, "Exit"):
                break
            
            if smtx_event == "Yes":
                if not nwtl:
                    print("There is no non-Wi-Fi Data.")
                else:
                    stuff_for_not_wlan = Not_WiFi(data = data, tech_list = nwtl, transmitter_names = "", exposure_conditions = "", reported_results_filepath = rwf, summary_results_filepath = swf, smtx_results_filepath = stxwf, log_dir = cwd)
                    
                    stuff_for_not_wlan.smtx_tech_results()
                
                if not wtl:
                    print("There is no non-Wi-Fi Data.")
                else:
                    stuff_for_wlan = Wlan(data = data, tech_list = wtl, transmitter_names = "", exposure_conditions = "", reported_results_filepath = rwf, summary_results_filepath = swf, smtx_results_filepath = stxwf, log_dir = cwd)
                    
                    stuff_for_wlan.smtx_tech_results()
                
                reported_excel(stxwf)
                
                sg.popup("Done", button_color = ("black", "#D3D3D3"), location = (center_window(primary_window, new_window = smtx_results)))
                
                smtx_results.close()
                
        smtx_results.close()
    
    else:
        sg.popup("The Smart Transmit directory has not been entered;\nplease enter the appropriate directory", button_color = ("black", "#D3D3D3"), location = (center_window(primary_window, new_window = smtx_results)))

def tech_band(primary_window, wb_1, rr_out, sr_out, sec_out, stx_out, rswf, sswf, sec1fccf, sec1isedf, stxf, cwd, data, exl):
    """
    Bulk of the program where it filters the raw SAR data, concatenates it into one dataframe,
    records the output directories for all the other sub-functions, and exposure conditions.
    
    Args:
        wb_1 (str): String filepath to a workbooks directory
        rr_out (str): String filepath for the reported results directory
        sr_out (str): String filepath for the summary results directory
        sec_out (str): String filepath for the Section 1 directory
        stx_out (str): String filepath for the TAS directory
        rswf (str): String filepath for the reported results
        sswf (str): String filepath for the summary results 
        sec1fccf (str): String filepath for Section 1 FCC
        sec1isedf (str): String filepath for Section 1 ISED
        stxf (str): String filepath for TAS
        cwd (str): String filepath for the current working directory
        data (dataframe): Concatenated dataframe of all the raw SAR data
        exl (list): List of applicable Exposure Conditions
    """
    
    tech_layout = [
        [
            sg.Text("Select each applicable technology and band that the product supports:", justification = "left")
        ],
        
        [
            sg.Button("GSM", key = "GSM", button_color = ("black", "#D3D3D3")),
            sg.Button("W-CDMA", key = "W-CDMA", button_color = ("black", "#D3D3D3")),
            sg.Button("LTE", key = "LTE", button_color = ("black", "#D3D3D3")),
            sg.Button("FR1", key = "FR1", button_color = ("black", "#D3D3D3")),
            sg.Button("WLAN", key = "WLAN", button_color = ("black", "#D3D3D3")),
            sg.Button("Bluetooth", key = "Bluetooth", button_color = ("black", "#D3D3D3")),
            sg.Button("Thread", key = "Thread", button_color = ("black", "#D3D3D3")),
            sg.Button("MSS/NTN", key = "MSS/NTN", button_color = ("black", "#D3D3D3"))
        ],
        
        [
            sg.Exit(button_color = ("white", "red")),
            sg.Button("Generate Reported Results Workbook", key = "-REPORTED_RESULTS-", button_color = ("black", "#D3D3D3")),
            sg.Button("Generate Summary Results Workbook", key = "-SUMMARY_RESULTS-", button_color = ("black", "#D3D3D3")),
            sg.Button("Generate Section 1 Workbook", key = "-SEC1_RESULTS-", button_color = ("black", "#D3D3D3")),
            sg.Button("Generate Smart Transmit Workbook", key = "-SMTX_RESULTS-", button_color = ("black", "#D3D3D3"))
        ]
    ]
    
    tech_window = sg.Window("Tech List", tech_layout, element_justification = "center", finalize = True)
    center_window(primary_window, new_window = tech_window)
    
    gsm_tech_list    = []
    wcdma_tech_list  = []
    lte_tech_list    = []
    fr1_tech_list    = []
    wlan_tech_list   = []
    bt_tech_list     = []
    thread_tech_list = []
    mss_tech_list    = []
    
    while True and wb_1 != "" and rr_out != "" or sr_out != "" or sec_out != "" or stx_out != "":
        tech_event, tech_values = tech_window.read()
        if tech_event in (sg.WINDOW_CLOSED, "Exit"):
            data = None
            break
        
        if tech_event == "GSM":
            try:
                gsm_tech_list = gsm_tech_window(primary_window = tech_window)
            except UnboundLocalError:
                pass
        
        if tech_event == "W-CDMA":
            try:
                wcdma_tech_list = wcdma_tech_window(primary_window = tech_window)
            except UnboundLocalError:
                pass
        
        if tech_event == "LTE":
            try:
                lte_tech_list = lte_tech_window(primary_window = tech_window)
            except UnboundLocalError:
                pass
        
        if tech_event == "FR1":
            try:
                fr1_tech_list = fr1_tech_window(primary_window = tech_window)
            except UnboundLocalError:
                pass
        
        if tech_event == "WLAN":
            try:
                wlan_tech_list = wlan_tech_window(primary_window = tech_window)
            except UnboundLocalError:
                pass
        
        if tech_event == "Bluetooth":
            try:
                bt_tech_list = bt_tech_window(primary_window = tech_window)
            except UnboundLocalError:
                pass
        
        if tech_event == "Thread":
            try:
                thread_tech_list = thread_tech_window(primary_window = tech_window)
            except UnboundLocalError:
                pass
        
        if tech_event == "MSS/NTN":
            try:
                mss_tech_list = mss_tech_window(primary_window = tech_window)
            except:
                pass
        
        if tech_event == "-REPORTED_RESULTS-":
            gsm    = gsm_tech_list
            wcdma  = wcdma_tech_list
            lte    = lte_tech_list
            fr1    = fr1_tech_list
            wlan   = wlan_tech_list
            bt     = bt_tech_list
            thread = thread_tech_list
            mss    = mss_tech_list
            
            collection = [gsm, wcdma, lte, fr1, bt, thread, mss]
            
            not_wifi_tech_list = [new for index, name in enumerate(collection) for new in name if new != ""]
            wifi_tech_list     = [tech for index, tech in enumerate(wlan)]
            
            reported_results(primary_window = tech_window, rr_out = rr_out, nwtl = not_wifi_tech_list, wtl = wifi_tech_list, data = data, rwf = rswf, swf = sswf, stxwf = stxf, cwd = cwd)
        
        if tech_event == "-SUMMARY_RESULTS-":
            gsm    = gsm_tech_list
            wcdma  = wcdma_tech_list
            lte    = lte_tech_list
            fr1    = fr1_tech_list
            wlan   = wlan_tech_list
            bt     = bt_tech_list
            thread = thread_tech_list
            mss    = mss_tech_list
            
            collection = [gsm, wcdma, lte, fr1, bt, thread, mss]
            
            not_wifi_tech_list = [new for index, name in enumerate(collection) for new in name if new != ""]
            wifi_tech_list     = [tech for index, tech in enumerate(wlan)]
            
            summary_results(primary_window = tech_window, sr_out = sr_out, nwtl = not_wifi_tech_list, wtl = wifi_tech_list, data = data, rwf = rswf, swf = sswf, stxwf = stxf, cwd = cwd, exl = exl)
        
        if tech_event == "-SEC1_RESULTS-":
            sec1_results(primary_window = tech_window, sec_out = sec_out, data = data, fcc = sec1fccf, ised = sec1isedf, cwd = cwd)
        
        if tech_event == "-SMTX_RESULTS-":
            gsm    = gsm_tech_list
            wcdma  = wcdma_tech_list
            lte    = lte_tech_list
            fr1    = fr1_tech_list
            wlan   = wlan_tech_list
            bt     = bt_tech_list
            thread = thread_tech_list
            mss    = mss_tech_list
            
            collection = [gsm, wcdma, lte, fr1, bt, thread, mss]
            
            not_wifi_tech_list = [new for index, name in enumerate(collection) for new in name if new != ""]
            wifi_tech_list     = [tech for index, tech in enumerate(wlan)]
            
            smtx_results(primary_window = tech_window, stx_out = stx_out, nwtl = not_wifi_tech_list, wtl = wifi_tech_list, data = data, rwf = rswf, swf = sswf, stxwf = stxf, cwd = cwd)
        
    else:
        if wb_1 == "":
            sg.popup("Missing the directory to Workbook 1; please enter the appropriate directory.", button_color = ("white", "red"), location = (center_window(primary_window, new_window = tech_window)))
        elif rr_out == "" and sr_out == "" and sec_out == "" and stx_out == "":
            sg.popup("Missing the directory to output the desired workbook; please enter the appropriate directory.", button_color = ("white", "red"), location = (center_window(primary_window, new_window = tech_window)))
            
    tech_window.close()

def main_window():
    """
    Main window where the user can select all the directories needed for file output
    """
    
    try:
        cwd = os.getcwd()
        
        sg.theme("Dark")
        
        file_list_column = [
            [
                sg.Text("Output Directory for Reported Table: "),
                sg.Input(key = "-REPORT_OUT-"),
                sg.FolderBrowse(key = "-REPORT_OUT_BROWSE-", button_color = ("black", "#D3D3D3"))
            ],
            
            [
                sg.Text("Output Directory for Summary Table: "),
                sg.Input(key = "-SUM_OUT-"),
                sg.FolderBrowse(key = "-SUM_OUT_BROWSE-", button_color = ("black", "#D3D3D3"))
            ],
            
            [
                sg.Text("Output Directory for Section 1 Table: "),
                sg.Input(key = "-SEC1_OUT-"),
                sg.FolderBrowse(key = "-SEC1_OUT_BROWSE-", button_color = ("black", "#D3D3D3"))
            ],
            
            [
                sg.Text("Output Directory for Smart Transmit Table: "),
                sg.Input(key = "-SMTX_OUT-"),
                sg.FolderBrowse(key = "-SMTX_OUT_BROWSE-", button_color = ("black", "#D3D3D3"))
            ],
        ]
        
        selected_files_column_1 = [
            [
                sg.Text("Workbook 01:"),
                sg.Input(key = "-Workbook_1-"),
                sg.FileBrowse(key = "-WB1_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 02:"),
                sg.Input(key = "-Workbook_2-"),
                sg.FileBrowse(key = "-WB2_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 03:"),
                sg.Input(key = "-Workbook_3-"),
                sg.FileBrowse(key = "-WB3_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 04:"),
                sg.Input(key = "-Workbook_4-"),
                sg.FileBrowse(key = "-WB4_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 05:"),
                sg.Input(key = "-Workbook_5-"),
                sg.FileBrowse(key = "-WB5_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 06:"),
                sg.Input(key = "-Workbook_6-"),
                sg.FileBrowse(key = "-WB6_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 07:"),
                sg.Input(key = "-Workbook_7-"),
                sg.FileBrowse(key = "-WB7_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 08:"),
                sg.Input(key = "-Workbook_8-"),
                sg.FileBrowse(key = "-WB8_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 09:"),
                sg.Input(key = "-Workbook_9-"),
                sg.FileBrowse(key = "-WB9_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 10:"),
                sg.Input(key = "-Workbook_10-"),
                sg.FileBrowse(key = "-WB10_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 11:"),
                sg.Input(key = "-Workbook_11-"),
                sg.FileBrowse(key = "-WB11_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
        ]
        
        selected_files_column_2 = [
            [
                sg.Text("Workbook 12:"),
                sg.Input(key = "-Workbook_12-"),
                sg.FileBrowse(key = "-WB12_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 13:"),
                sg.Input(key = "-Workbook_13-"),
                sg.FileBrowse(key = "-WB13_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 14:"),
                sg.Input(key = "-Workbook_14-"),
                sg.FileBrowse(key = "-WB14_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 15:"),
                sg.Input(key = "-Workbook_15-"),
                sg.FileBrowse(key = "-WB15_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 16:"),
                sg.Input(key = "-Workbook_16-"),
                sg.FileBrowse(key = "-WB16_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 17:"),
                sg.Input(key = "-Workbook_17-"),
                sg.FileBrowse(key = "-WB17_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 18:"),
                sg.Input(key = "-Workbook_18-"),
                sg.FileBrowse(key = "-WB18_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 19:"),
                sg.Input(key = "-Workbook_19-"),
                sg.FileBrowse(key = "-WB19_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 20:"),
                sg.Input(key = "-Workbook_20-"),
                sg.FileBrowse(key = "-WB20_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 21:"),
                sg.Input(key = "-Workbook_21-"),
                sg.FileBrowse(key = "-WB21_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
            
            [
                sg.Text("Workbook 22:"),
                sg.Input(key = "-Workbook_22-"),
                sg.FileBrowse(key = "-WB22_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
            ],
        ]
        
        # selected_files_column_3 = [
        #     [
        #         sg.Text("Workbook 23:"),
        #         sg.Input(key = "-Workbook_23-"),
        #         sg.FileBrowse(key = "-WB23_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
            
        #     [
        #         sg.Text("Workbook 24:"),
        #         sg.Input(key = "-Workbook_24-"),
        #         sg.FileBrowse(key = "-WB24_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
            
        #     [
        #         sg.Text("Workbook 25:"),
        #         sg.Input(key = "-Workbook_25-"),
        #         sg.FileBrowse(key = "-WB25_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
            
        #     [
        #         sg.Text("Workbook 26:"),
        #         sg.Input(key = "-Workbook_26-"),
        #         sg.FileBrowse(key = "-WB26_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
            
        #     [
        #         sg.Text("Workbook 27:"),
        #         sg.Input(key = "-Workbook_27-"),
        #         sg.FileBrowse(key = "-WB27_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
            
        #     [
        #         sg.Text("Workbook 28:"),
        #         sg.Input(key = "-Workbook_28-"),
        #         sg.FileBrowse(key = "-WB28_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
            
        #     [
        #         sg.Text("Workbook 29:"),
        #         sg.Input(key = "-Workbook_29-"),
        #         sg.FileBrowse(key = "-WB29_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
            
        #     [
        #         sg.Text("Workbook 30:"),
        #         sg.Input(key = "-Workbook_30-"),
        #         sg.FileBrowse(key = "-WB30_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
            
        #     [
        #         sg.Text("Workbook 31:"),
        #         sg.Input(key = "-Workbook_31-"),
        #         sg.FileBrowse(key = "-WB31_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
            
        #     [
        #         sg.Text("Workbook 32:"),
        #         sg.Input(key = "-Workbook_32-"),
        #         sg.FileBrowse(key = "-WB32_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
            
        #     [
        #         sg.Text("Workbook 33:"),
        #         sg.Input(key = "-Workbook_33-"),
        #         sg.FileBrowse(key = "-WB33_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        #     ],
        # ]
        
        initial_layout = [
            [
                sg.Column(file_list_column, element_justification = "right"),
                sg.VerticalSeparator(),
                sg.Column(selected_files_column_1, element_justification = "left"),
                sg.VerticalSeparator(),
                sg.Column(selected_files_column_2, element_justification = "left"),
            ],
            
            [
                sg.Exit(button_color = ("white", "red")),
                sg.Button("Clear", key = "-CLEAR-", button_color = ("black", "#D3D3D3")),
                sg.Button("Technologies & Bands", key = "-TECH_BAND-", button_color = ("black", "#D3D3D3")),
                sg.Button("Summary Workbook Comparison", key = "-SUM_COMPARE-", button_color = ("black", "#D3D3D3")),
            ]
        ]
        
        initial_window = sg.Window("Reported & Summary Table Generator", initial_layout, finalize = True)
        
        while True:
            event, values = initial_window.read()
            # print(f"Event: {event}; Value: {values}")
            if event in (sg.WINDOW_CLOSED, "Exit"):
                break
            
            if event == "-CLEAR-":
                keys_to_clear = ["-REPORT_OUT-", "-SUM_OUT-", "-SEC1_OUT-", "-SMTX_OUT-", "-Workbook_1-", "-Workbook_2-", "-Workbook_3-", "-Workbook_4-", "-Workbook_5-", "-Workbook_6-", "-Workbook_7-", "-Workbook_8-", "-Workbook_9-", "-Workbook_10-", "-Workbook_11-", "-Workbook_12-", "-Workbook_13-", "-Workbook_14-", "-Workbook_15-", "-Workbook_16-", "-Workbook_17-", "-Workbook_18-", "-Workbook_19-", "-Workbook_20-", "-Workbook_21-", "-Workbook_22-"]
                
                for key in keys_to_clear:
                    initial_window[key].update("")
                    values[key] = ""
            
            if values["-REPORT_OUT-"] != "":        
                reported_sar_workbook_filename = parse_reported_workbook_name(values["-Workbook_1-"])
                reported_sar_workbook_directory = os.path.normpath(values["-REPORT_OUT_BROWSE-"])
                reported_sar_workbook_filepath = os.path.join(reported_sar_workbook_directory, reported_sar_workbook_filename) # reported_sar_workbook_directory + "/" + reported_sar_workbook_filename
                
                # print(reported_sar_workbook_filepath)
            else:
                reported_sar_workbook_filepath = ""
            
            if values["-SUM_OUT-"] != "":
                worst_case_summary_sar_workbook_filename = parse_summary_workbook_name(values["-Workbook_1-"])
                worst_case_summary_sar_workbook_directory = os.path.normpath(values["-SUM_OUT_BROWSE-"])
                worst_case_summary_sar_workbook_filepath = os.path.join(worst_case_summary_sar_workbook_directory, worst_case_summary_sar_workbook_filename) # worst_case_summary_sar_workbook_directory + "/" + worst_case_summary_sar_workbook_filename
                
                # print(worst_case_summary_sar_workbook_filepath)
            else:
                worst_case_summary_sar_workbook_filepath = ""
            
            if values["-SEC1_OUT-"] != "":
                sec1_fcc_workbook_filename = parse_sec1_fcc_workbook_name(values["-Workbook_1-"])
                sec1_fcc_workbook_directory = os.path.normpath(values["-SEC1_OUT_BROWSE-"])
                sec1_fcc_workbook_filepath = os.path.join(sec1_fcc_workbook_directory, sec1_fcc_workbook_filename) # sec1_fcc_workbook_directory + "/" + sec1_fcc_workbook_filename
                
                sec1_ised_workbook_filename = parse_sec1_ised_workbook_name(values["-Workbook_1-"])
                sec1_ised_workbook_directory = os.path.normpath(values["-SEC1_OUT_BROWSE-"])
                sec1_ised_workbook_filepath = os.path.join(sec1_ised_workbook_directory, sec1_ised_workbook_filename) # sec1_ised_workbook_directory + "/" + sec1_ised_workbook_filename
                
                # print(worst_case_summary_sar_workbook_filepath)
            else:
                sec1_fcc_workbook_filepath = ""
                sec1_ised_workbook_filepath = ""
            
            if values["-SMTX_OUT-"] != "":
                smtx_sar_workbook_filename = parse_smtx_workbook_name(values["-Workbook_1-"])
                smtx_sar_workbook_directory = os.path.normpath(values["-SMTX_OUT_BROWSE-"])
                smtx_sar_workbook_filepath = os.path.join(smtx_sar_workbook_directory, smtx_sar_workbook_filename) # smtx_sar_workbook_directory + "/" + smtx_sar_workbook_filename
                
                # print(worst_case_summary_sar_workbook_filepath)
            else:
                smtx_sar_workbook_filepath = ""
            
            if event == "-SUM_COMPARE-":
                summary_compare(cwd)
                
            if values["-Workbook_1-"] != "":
                rem_key = ["-REPORT_OUT-", "-REPORT_OUT_BROWSE-", "-SUM_OUT-", "-SUM_OUT_BROWSE-", "-SEC1_OUT-", "-SEC1_OUT_BROWSE-", "-SMTX_OUT-", "-SMTX_OUT_BROWSE-" "-WB1_IN_BROWSE-", "-WB2_IN_BROWSE-", "-WB3_IN_BROWSE-", "-WB4_IN_BROWSE-", "-WB5_IN_BROWSE-", "-WB6_IN_BROWSE-", "-WB7_IN_BROWSE-", "-WB8_IN_BROWSE-", "-WB9_IN_BROWSE-", "-WB10_IN_BROWSE-", "-WB11_IN_BROWSE-", "-WB12_IN_BROWSE-", "-WB13_IN_BROWSE-", "-WB14_IN_BROWSE-", "-WB15_IN_BROWSE-", "-WB16_IN_BROWSE-", "-WB17_IN_BROWSE-", "-WB18_IN_BROWSE-", "-WB19_IN_BROWSE-", "-WB20_IN_BROWSE-", "-WB21_IN_BROWSE-", "-WB22_IN_BROWSE-"]
                
                wbs_dict = {key: values[key] for key in values if key not in rem_key}
                
                wbs_with_duplicates = list(wbs_dict.values())
                
                rem_empty = [name for index, name in enumerate(wbs_with_duplicates) if name != ""]
                
                wbs = [name for index, name in enumerate(rem_empty) if name not in rem_empty[:index]]
                
                data = [pd.read_excel(wbs[index], sheet_name = "Data") for index, name in enumerate(wbs)]
                
                expose_list = ["Head", "Body-worn", "Body & Hotspot", "Hotspot", "Extremity"]
                
                # print(wbs)
            
            if event == "-TECH_BAND-":
                try:
                    tech_band(primary_window = initial_window, wb_1 = values["-Workbook_1-"], rr_out = values["-REPORT_OUT-"], sr_out = values["-SUM_OUT-"], sec_out = values["-SEC1_OUT-"], stx_out = values["-SMTX_OUT-"], rswf = reported_sar_workbook_filepath, sswf = worst_case_summary_sar_workbook_filepath, sec1fccf = sec1_fcc_workbook_filepath, sec1isedf = sec1_ised_workbook_filepath, stxf = smtx_sar_workbook_filepath, cwd = cwd, data = data, exl = expose_list)
                except UnboundLocalError:
                    window_x_local, window_y_local = initial_window.current_location()
                    window_width, window_height = initial_window.size
                    
                    sg.popup("Either nothing has been entered or a workbook directory is missing, please try again", title = "Missing Information", button_color = ("white", "red"), location = ((window_width // 2) + window_x_local, (window_height // 2) + window_y_local))
                    pass
        
        initial_window.close()
    
    except Exception as e:
        cwd_e = os.getcwd()
        error_log = os.path.join(cwd_e, "error.log")
        
        logger = logging.getLogger("SAR GUI")
        logger.setLevel(logging.ERROR)
        
        lfh = logging.FileHandler(error_log)
        lfh.setLevel(logging.ERROR)
        
        formatter = logging.Formatter("\n%(asctime)s - %(name)s - %(levelname)s - %(message)s")
        lfh.setFormatter(formatter)
        
        logger.addHandler(lfh)
        
        logger.exception(e)
        
        sg.popup_error_with_traceback(f"An error has occured, here is the info:\n", e, f"\nMore information can be found here: {error_log}")

if __name__ == "__main__":
    main_window()